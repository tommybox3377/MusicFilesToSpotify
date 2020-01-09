from openpyxl import Workbook, load_workbook
from tinytag import TinyTag
import re
import Creds
import spotipy
import os
import spotipy.util as util

# Private Variables
main_music_dir = r"C:\Users\..."
playlist_url = "ABC123"

# Options
delete_album_art = False

music_files = []
paths_of_album_art = []
unsupported_file_type = []
scope = 'playlist-modify-public user-library-modify playlist-modify-private'
os.environ["SPOTIPY_CLIENT_ID"] = Creds.ClientID
os.environ["SPOTIPY_CLIENT_SECRET"] = Creds.SecretID
os.environ["SPOTIPY_REDIRECT_URI"] = "http://localhost/"


def get_list_of_music_paths():
    for r, d, f in os.walk(main_music_dir):
        for file in f:
            if '.mp3' in file.lower() or '.wma' in file or '.m4a' in file:
                music_files.append(os.path.join(r, file))
            elif '.jpg' in file:
                paths_of_album_art.append(os.path.join(r, file))
            else:
                unsupported_file_type.append(os.path.join(r, file))


def dlt_album_art():
    if delete_album_art:
        for i in paths_of_album_art:
            os.remove(i)


def create_xl():
    workbook = Workbook()
    workbook.create_sheet("All Music")
    workbook.create_sheet("Unsupported Files")
    workbook.create_sheet("Transferred To Spotify")
    workbook.create_sheet("Could Not Find On Spotify")
    workbook.create_sheet("Failed Adding to Spotify")
    del workbook["Sheet"]
    sheet = workbook["All Music"]
    sheet["A1"] = "Full File Name"
    sheet["B1"] = "Album"
    sheet["C1"] = "Artist"
    sheet["D1"] = "Title"
    sheet["E1"] = "Spotify ID"
    sheet["F1"] = "Transfered to Spotify"
    sheet["G1"] = "Year"
    sheet["H1"] = "File Name"
    sheet["I1"] = "Song URL"
    workbook.save(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")


def get_song_metadata():
    workbook = load_workbook(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")
    sheet = workbook["All Music"]
    for music_file in music_files:
        try:
            tag = TinyTag.get(music_file)
            sheet.append((music_file, tag.album, tag.artist, tag.title, "", False, tag.year, os.path.basename(music_file)[:-4]))
        except:
            unsupported_file_type.append(music_file)
            print(f"{music_file} could not be processed, added to unsupported file log")
    workbook.save(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")


def log_unsupported_file_types():
    workbook = load_workbook(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")
    sheet = workbook["Unsupported Files"]
    for file in unsupported_file_type:
        sheet.append((file,))
    workbook.save(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")


def clean_title(title):
    t = re.sub(r"\(.*?\)", "", title)
    if t:
        return t.strip()
    else:
        return title


def clean_file_name(name):
    if name:
        n = re.sub("[^0-9a-zA-Z ]", '', name)
        if n:
            return n
        else:
            return name
    else:
        return name


def prompt_for_token():
    util.prompt_for_user_token(Creds.spotfiy_username, scope)


def get_token():
    token = util.prompt_for_user_token(Creds.spotfiy_username, scope)
    if token:
        return token
    else:
        print("Failed at receiving token")


def create_spotify_query(metadata, albm=True, filename=False):
    album = metadata[1] if metadata[1] else None
    artist = metadata[2] if metadata[2] else None
    track = metadata[3] if metadata[3] else None
    q_list = []
    if filename:
        return clean_file_name(metadata[7])
    if artist and track and albm:
        if artist:
            q_list.append(artist)
        if track:
            q_list.append(clean_title(track))
        if album:
            q_list.append(album)
        q_string = " ".join(q_list)
        return q_string
    elif artist and track:
        if artist:
            q_list.append(artist)
        if track:
            q_list.append(clean_title(track))
        q_string = " ".join(q_list)
        return q_string
    else:
        return clean_file_name(metadata[7])


def find_spotify_ids():
    workbook = load_workbook(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")
    sheet = workbook["All Music"]
    spotify = spotipy.Spotify(auth=get_token())
    for i, row in enumerate(sheet.values):
        if i > 38000:
            query = create_spotify_query(row)
            query = query[0:300] if query else query
            if query:
                results = spotify.search(q=query, type='track', limit=1)
                if len(results["tracks"]["items"]) > 0:
                    sheet["E" + str(i+1)] = results["tracks"]["items"][0]["id"]
                    sheet["I" + str(i+1)] = results["tracks"]["items"][0]["external_urls"]["spotify"]
                else:
                    query = create_spotify_query(row, albm=False)
                    query = query[0:300] if query else query
                    if query:
                        results = spotify.search(q=query, type='track', limit=1)
                        if len(results["tracks"]["items"]) > 0:
                            sheet["E" + str(i + 1)] = results["tracks"]["items"][0]["id"]
                            sheet["I" + str(i + 1)] = results["tracks"]["items"][0]["external_urls"]["spotify"]
                        else:
                            query = create_spotify_query(row, filename=True)
                            query = query[0:300] if query else query
                            if query:
                                results = spotify.search(q=query, type='track', limit=1)
                                if len(results["tracks"]["items"]) > 0:
                                    sheet["E" + str(i + 1)] = results["tracks"]["items"][0]["id"]
                                    sheet["I" + str(i + 1)] = results["tracks"]["items"][0]["external_urls"]["spotify"]
                                else:
                                    sheet["E" + str(i + 1)] = "no match found"
                                    sheet["I" + str(i + 1)] = query
            if i % 1000 == 0:
                workbook.save(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")
                print(i)
    workbook.save(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")


def add_music_to_playlist():
    find_spotify_ids()
    workbook = load_workbook(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")
    sheet = workbook["All Music"]
    data = sheet["E"]
    ids = set()
    for datum in data:
        if datum.value != "no match found" and datum.value != "Spotify ID" and datum.value:
            ids.add(datum.value)
    print("adding")
    spotify = spotipy.Spotify(auth=get_token())
    spotify.trace = False
    for id in ids:
        try:
            track = (id,)
            print(track)
            spotify.user_playlist_add_tracks(Creds.spotfiy_username, playlist_url, track)
        except:
            workbook = load_workbook(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")
            sheet2 = workbook["Failed Adding to Spotify"]
            sheet2.append((id,))
            workbook.save(filename=main_music_dir + "\MusicFilesToSpotifyLog.xlsx")
            spotify = spotipy.Spotify(auth=get_token())
            print(f"failed adding {id}")


prompt_for_token()
get_list_of_music_paths()
create_xl()
get_song_metadata()
log_unsupported_file_types()
add_music_to_playlist()
if delete_album_art:
    dlt_album_art()